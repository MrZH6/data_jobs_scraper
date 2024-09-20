import csv
import re
import json
import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import time

# TODO: Jobs.cz:
#  - přidat města
#  - procházet všechny nabídky ??
# TODO: Welcome to The Jungle:
#  - location beru pouze první město
#  - odebrat query parametr a procházet všechny pozice
#  - doplnit plat
# TODO: headers do dictionary - {col: name, col_width}
# TODO: sjednotit datum inzerátu


start_time = time.time()

PATTERN_1 = r"\b(BI|ML|AI)\b"
PATTERN_2 = r"(?i)(data|\bdat\b|datov|reporting|business intelligence|machine learning|uměl.{1} inteligenc.{1}|strojov.{1,3} učení)"
HEADERS = ["Portál", "Číslo", "Název pozice", "Link", "Firma", "Lokace", "Typ úvazku", "Datum inzerátu", "Tag", "Plat"]

def fetch_jobs_data(url):
    try:
        response = requests.get(url)
        response.raise_for_status()  # Check if the request was successful
    except requests.exceptions.RequestException as e:
        print(f"Error fetching the URL: {e}")
        return None

    return response.text

def jobs():
    job_list, job_count = [], 0
    max_pages = 100
    cities = ["praha", "brno", "ostrava", "plzen"]

    for city in cities:

        for page in range(1, max_pages+1):
            query = "q%5B%5D=data%20engineer&q%5B%5D=data%20analyst&q%5B%5D=BI%20developer&q%5B%5D=ML%20Engineer&q%5B%5D=Data%20Scientist"
            locality_radius = "locality%5Bradius%5D=30"
            url = f"https://www.jobs.cz/prace/{city}/?{query}&page={page}&{locality_radius}"
            html_content = fetch_jobs_data(url)

            if html_content:
                soup = BeautifulSoup(html_content, 'html.parser')
                page_overflow = soup.find('div', {'data-test': 'page-overflow-alert'})

                if page_overflow:
                    print(f"Max page is {page-1}")
                    break

                for job in soup.find_all('article', class_='SearchResultCard'):

                    title = job.find('header', class_='SearchResultCard__header').find('h2', class_="SearchResultCard__title").get_text(strip=True)

                    company = job.find('footer', class_='SearchResultCard__footer').find('li', class_="SearchResultCard__footerItem").get_text(strip=True)

                    location = job.find('footer', class_='SearchResultCard__footer').find('li', {"data-test" : "serp-locality"}).get_text(strip=True)

                    job_url = job.find('header', class_='SearchResultCard__header').find('a', class_='link-primary SearchResultCard__titleLink').get('href')

                    insert_date = job.find('header', class_='SearchResultCard__header').find('div', class_="SearchResultCard__status").get_text(strip=True)

                    pay_tag = job.find('div', class_='SearchResultCard__body').find('span', class_='Tag Tag--success Tag--small Tag--subtle')
                    pay = pay_tag.get_text(strip=True) if pay_tag else None

                    if insert_date == "Příležitost dne":
                        tag = "HOT"
                    elif insert_date == "Přidáno dnes" or insert_date == "Přidáno včera":
                        tag = "NEW"
                    else:
                        tag = None

                    if re.search(PATTERN_1, title) or re.search(PATTERN_2, title):
                        print(f"{title} | {company}")
                        job_count += 1
                        # portal, job_count, title, job_url, company, location, shifts, insert_date, tag, pay
                        job_list.append(["jobs.cz", job_count, title, job_url, company, location, None, insert_date, tag, pay])
            else:
                print(f"Failed to retrieve data from city {city} at page {page}")

    print(f"jobs.cz jobs found: {job_count}")
    return job_list

def startupjobs():
    job_list, job_count = [], 0
    url = "https://www.startupjobs.cz/api/offers"
    page = 1

    while True:
        print(f"Page: {page}")
        response = requests.request("GET", url, params={"page": page})
        data = response.json()
        for job in data["resultSet"]:
            title = job["name"]
            company = job["company"]
            location = job["locations"]
            job_url = "https://www.startupjobs.cz" + job["url"]
            shifts = job["shifts"]
            tag = "HOT" if job["isHot"] else None

            if job["salary"]:
                salary_min, salary_max = job["salary"]["min"], job["salary"]["max"]
                salary_currency, salary_measure = job["salary"]["currency"], job["salary"]["measure"]
                pay = f"{salary_min} - {salary_max} {salary_currency} {salary_measure}"
            else:
                pay = None

            if re.search(PATTERN_1, title) or re.search(PATTERN_2, title):
                job_count += 1
                # portal, job_count, title, job_url, company, location, shifts, insert_date, tag, pay
                job_list.append(["startupjobs.cz", job_count, title, job_url, company, location, shifts, None, tag, pay])
                print(f"{title} | {company}")

        max_page = data["paginator"]["max"]

        if page == max_page:
            break

        page += 1

    print(f"startupjobs.cz jobs found: {job_count}")
    return job_list

def cocuma():
    job_list, job_count = [], 0
    max_pages = 100

    for page in range(1, max_pages+1):
        print(page)
        if page == 1:
            url = "https://www.cocuma.cz/jobs"
        else:
            url = f"https://www.cocuma.cz/jobs/page/{page}"
        html_content = fetch_jobs_data(url)

        if html_content:
            soup = BeautifulSoup(html_content, 'html.parser')

            for job in soup.find_all('div', class_='col-md-6 col-lg-4'):
                if job.find('a', class_='job-thumbnail'):

                    title_tag = job.find('a', class_='job-thumbnail').find('p', class_='job-thumbnail-title')
                    title = title_tag.get_text(strip=True) if title_tag else None

                    company_tag = job.find('a', class_='job-thumbnail').find('p', class_='job-thumbnail-company')
                    company = company_tag.get_text(strip=True) if company_tag else None

                    location_tag = job.find('a', class_='job-thumbnail').find('p', class_='job-thumbnail-city')
                    location = location_tag.get_text(strip=True) if location_tag else None

                    tag_tag = job.find('a', class_='job-thumbnail').find('div', class_='job-thumbnail-badge')
                    tag = tag_tag.get_text(strip=True) if tag_tag else None

                    shifts_tag = job.find('a', class_='job-thumbnail').find('p', class_='job-thumbnail-work-shedule')
                    shifts = shifts_tag.get_text(strip=True) if shifts_tag else None

                    job_url = "https://www.cocuma.cz" + job.find('a', class_='job-thumbnail').get('href')

                    if title and (re.search(PATTERN_1, title) or re.search(PATTERN_2, title)):
                        job_count += 1
                        # portal, job_count, title, job_url, company, location, shifts, insert_date, tag, pay
                        job_list.append(["cocuma.cz", job_count, title, job_url, company, location, shifts, None, tag, None])
                        print(f"{title} | {company}")

        else:
            print(f"Failed to retrieve data from page {page}")
            break

    print(f"cocuma.cz jobs found: {job_count}")
    return job_list

def futureproof():
    job_list, job_count = [], 0
    url = "https://jobs.fproof.eu/recruit/v2/public/Job_Openings?pagename=Careers&source=CareerSite"

    response = requests.request("GET", url)
    data = response.json()
    for job in data["data"]:
        title = job["Posting_Title"]
        location = job["City"]
        job_url = job["$url"]
        shifts = job["Job_Type"]
        insert_date = job["Date_Opened"]

        if title and (re.search(PATTERN_1, title) or re.search(PATTERN_2, title)):
            job_count += 1
            # portal, job_count, title, job_url, company, location, shifts, insert_date, tag, pay
            job_list.append(["fproof.eu", job_count, title, job_url, None, location, shifts, insert_date, None, None])
            print(f"{title}")

    print(f"fproof.eu jobs found: {job_count}")
    return job_list

def welcome_to_the_jungle(api_key, application_id):
    job_list, job_count = [], 0
    url = "https://csekhvms53-dsn.algolia.net/1/indexes/*/queries?search_origin=job_search_client"
    max_pages = 50
    headers = {
        'X-Algolia-Api-Key': api_key,
        'X-Algolia-Application-Id': application_id,
        'Referer': 'https://www.welcometothejungle.com/',
        'Content-Type': 'application/json'
    }

    for page in range(0, max_pages+1):
        payload = json.dumps({
            "requests": [
                {
                    "indexName": "wttj_jobs_production_cs",
                    "params": f"filters=(%22offices.country_code%22%3A%22CZ%22)&page={page}&query=data"
                }
            ]
        })

        response = requests.request("POST", url, headers=headers, data=payload, timeout=10)
        response_json = response.json()
        response_jobs = response_json["results"][0]["hits"]

        print(f"Page {page}: {len(response_jobs)} jobs.")

        if len(response_jobs) == 0:
            break

        for job in response_jobs:
            title = job["name"]
            company = job["organization"]["name"]
            location = job["offices"][0]["city"]
            job_url = f"https://www.welcometothejungle.com/cs/companies/{job['organization']['slug']}/jobs/{job['slug']}"
            shifts = job["contract_type"]
            insert_date = job["published_at"]
            tag = "Boosted" if job["is_boosted"] else None

            if title and (re.search(PATTERN_1, title) or re.search(PATTERN_2, title)):
                job_count += 1
                # portal, job_count, title, job_url, company, location, shifts, insert_date, tag, pay
                job_list.append(["welcometothejungle.com", job_count, title, job_url, company, location, shifts, insert_date, tag, None])
                print(f"{title} | {company}")

    print(f"welcometothejungle.com jobs found: {job_count}")
    return job_list

# load credentials
with open("credentials.json", 'r') as file:
    credentials = json.load(file)

# fetch all job positions
jobs_all = jobs()
jobs_all += startupjobs()
jobs_all += cocuma()
jobs_all += futureproof()
jobs_all += welcome_to_the_jungle(credentials["welcome_to_the_jungle"]["api_key"], credentials["welcome_to_the_jungle"]["application_id"])

with open('output/jobs.csv', mode='w', newline='', encoding='utf-8') as file:
    writer = csv.writer(file)
    writer.writerow(HEADERS)
    writer.writerows(jobs_all)

df_jobs = pd.DataFrame(jobs_all, columns=HEADERS)
df_jobs.to_excel('output/jobs.xlsx', index=False, sheet_name='DATA_JOBS')

# modify Excel file
wb = load_workbook('output/jobs.xlsx')
ws = wb['DATA_JOBS']

ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 7
ws.column_dimensions["C"].width = 55
ws.column_dimensions["D"].width = 7
ws.column_dimensions["E"].width = 25
ws.column_dimensions["F"].width = 20
ws.column_dimensions["G"].width = 15
ws.column_dimensions["H"].width = 16
ws.column_dimensions["I"].width = 10
ws.column_dimensions["J"].width = 20

for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="808b96")

link_col = "D"
for row in range(2, ws.max_row + 1):
    cell = ws[f'{link_col}{row}']
    if cell.value.startswith('http'):
        ws[f'{link_col}{row}'].hyperlink = cell.value
        ws[f'{link_col}{row}'].value = "link"
        ws[f'{link_col}{row}'].style = "Hyperlink"

wb.save('output/jobs.xlsx')

end_time = time.time()
print(f"Execution time: {round(end_time - start_time, 1)} seconds")
